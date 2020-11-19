import os
import torch
import torch.nn as nn
import numpy as np
import termcolor
import matplotlib.pyplot as plt
from torch.utils.data import DataLoader
from torch import optim
from openpyxl import load_workbook

from BBNfolder.BBN_datasampler import COP_Data
from ReadExcel import get_valid_index
from BBNfolder.BBN import block_resnet_shared, block_resnet_separated, classifier


def toGreen(str): return termcolor.colored(str, "green", attrs=["bold"])
def toCyan(str): return termcolor.colored(str, "cyan", attrs=["bold"])
def toBlue(str): return termcolor.colored(str, "blue", attrs=["bold"])


def get_index(fname):
    with open(fname, 'r') as f:
        indexes = f.readlines()
        indexes = list(map(lambda x: int(x.replace('\n', '')), indexes))
    return np.asarray(indexes)


def train(ep, max_ep, models, dataset, criterion, optimizer, scheduler, device, fold_num, sample_num, batch_size):
    for m in models:
        m.train()
    ep_loss = 0.0
    pp = int(len(dataset) / 10)

    # generate random indexes for uniform sampler
    initial_ind = np.asarray(list(range(len(dataset))))
    uni_ind = np.random.choice(initial_ind, len(initial_ind), replace=False)
    imb_ratio = (sample_num[0] / sample_num[1]) / (sample_num[0] / sample_num[1] + 1)

    alpha = 1 - ((ep - 1) / max_ep) ** 2
    # softmax = nn.Softmax(dim=1)
    if len(dataset) % batch_size:
        iter_len = len(dataset) // batch_size + 1
    else:
        iter_len = len(dataset) // batch_size

    for i in range(iter_len):
        # stack for mini-batch
        input_a, input_b = torch.tensor([]).double().cuda(), torch.tensor([]).double().cuda()
        label_a, label_b = torch.tensor([]).long().cuda(), torch.tensor([]).long().cuda()
        for j in range(batch_size):
            sample_ind = i * batch_size + j
            if sample_ind == len(dataset):
                break
            class_choice = np.random.choice(2, 1, p=[1 - imb_ratio, imb_ratio])
            # uniform sampling and reversed sampling from dataset
            input1, label1 = dataset[uni_ind[sample_ind]]
            input2, label2 = dataset[int(initial_ind[np.random.choice(sample_num[int(not class_choice)], 1)
                                                     + class_choice * sample_num[1]])]
            input1, input2, label1, label2 = \
                input1.to(device), input2.to(device), \
                torch.tensor(label1[0]).to(device), torch.tensor(label2[0]).to(device)
            input_a, input_b = torch.cat([input_a, input1.unsqueeze(0)], dim=0),\
                               torch.cat([input_b, input2.unsqueeze(0)], dim=0)
            label_a, label_b = torch.cat([label_a, label1.unsqueeze(0)], dim=0),\
                               torch.cat([label_b, label2.unsqueeze(0)], dim=0)
        optimizer.zero_grad()
        # forward pass
        # (model_shared, model_sep1, model_sep2, model_classifier1, model_classifier2)
        (model_shared, model_sep1, model_sep2, model_classifier1, model_classifier2) = models

        output1 = model_shared(input_a.float())
        output1 = model_sep1(output1).flatten(1)
        output1 = model_classifier1(output1 * alpha)

        output2 = model_shared(input_b.float())
        output2 = model_sep2(output2).flatten(1)
        output2 = model_classifier2(output2 * (1 - alpha))

        output = output1 + output2

        loss = alpha * criterion(output, label_a.float().unsqueeze(1)) + \
               (1 - alpha) * criterion(output, label_b.float().unsqueeze(1))
        ep_loss += loss.item()
        loss.backward()
        optimizer.step()
        if scheduler is not None:
            scheduler.step(ep + i / len(dataset))
        if (i + 1) % pp == 0:
            print("Epoch: {}\tIter: {}/{}\tLoss: {}\tlr: {}\t{}-th fold loop".format(
                ep, i + 1, len(dataset), loss.item(), optimizer.param_groups[0]['lr'], fold_num))
    ep_loss = ep_loss / len(dataset)
    return models, ep_loss


def test(epoch, models, dataset, device, sample_num, batch_size):
    for m in models:
        m.test()
    sig = nn.Sigmoid()
    alpha = 0.5
    initial_ind = np.asarray(list(range(len(dataset))))
    uni_ind = np.random.choice(initial_ind, len(initial_ind), replace=False)
    imb_ratio = (sample_num[0] / sample_num[1]) / (sample_num[0] / sample_num[1] + 1)
    pp = int(len(dataset) / 10)
    num_correct = 0.0
    num_sample = 0.0
    zeros = np.zeros(2)
    ones = np.zeros(2)
    # softmax = nn.Softmax(dim=1)
    if len(dataset) % batch_size:
        iter_len = len(dataset) // batch_size + 1
    else:
        iter_len = len(dataset) // batch_size

    for i in range(iter_len):
        # stack for mini-batch
        input_a, input_b = torch.tensor([]).double().cuda(), torch.tensor([]).double().cuda()
        label_a, label_b = torch.tensor([]).long().cuda(), torch.tensor([]).long().cuda()
        for j in range(batch_size):
            sample_ind = i * batch_size + j
            if sample_ind == len(dataset):
                break
            class_choice = np.random.choice(2, 1, p=[1 - imb_ratio, imb_ratio])
            # uniform sampling and reversed sampling from dataset
            input1, label1 = dataset[uni_ind[sample_ind]]
            input2, label2 = dataset[int(initial_ind[np.random.choice(sample_num[int(not class_choice)], 1)
                                                     + class_choice * sample_num[1]])]
            input1, input2, label1, label2 = \
                input1.to(device), input2.to(device), \
                torch.tensor(label1[0]).to(device), torch.tensor(label2[0]).to(device)
            input_a, input_b = torch.cat([input_a, input1.unsqueeze(0)], dim=0), \
                               torch.cat([input_b, input2.unsqueeze(0)], dim=0)
            label_a, label_b = torch.cat([label_a, label1.unsqueeze(0)], dim=0), \
                               torch.cat([label_b, label2.unsqueeze(0)], dim=0)

        (model_shared, model_sep1, model_sep2, model_classifier1, model_classifier2) = models

        if len(label_a) == 1:
            break
        output1 = model_shared(input_a.float())
        output1 = model_sep1(output1).flatten(1)
        output1 = model_classifier1(output1 * alpha)

        output2 = model_shared(input_b.float())
        output2 = model_sep2(output2).flatten(1)
        output2 = model_classifier2(output2 * (1 - alpha))

        output = sig(output1 + output2)

        output[output < 0.5] = 0
        output[output >= 0.5] = 1

        corr_mask1 = output == label_a.unsqueeze(1)
        corr_mask2 = output == label_b.unsqueeze(1)
        num_correct += (alpha * corr_mask1.sum() + (1 - alpha) * corr_mask2.sum())
        num_sample += output.numel()
        ones[0] += np.asarray([alpha * float(cm1 and l1) + (1 - alpha) * float(cm2 and l2)
                               for cm1, cm2, l1, l2 in zip(corr_mask1, corr_mask2, label_a, label_b)]).sum()
        ones[1] += alpha * float((label_a == 1).sum()) + (1 - alpha) * float((label_b == 1).sum())
        zeros[0] += np.asarray([alpha * float(cm1 and not l1) + (1 - alpha) * float(cm2 and not l2)
                                for cm1, cm2, l1, l2 in zip(corr_mask1, corr_mask2, label_a, label_b)]).sum()
        zeros[1] += alpha * float((label_a == 0).sum()) + (1 - alpha) * float((label_b == 0).sum())
        if (i + 1) % pp == 0:
            print("Epoch: {}\tSample: {}/{}".format(epoch, i + 1, len(dataset)))
    avg_acc = (ones[0] / ones[1] + zeros[0] / zeros[1]) * 0.5
    total_acc = num_correct / num_sample

    return avg_acc, total_acc, zeros[0] / zeros[1], ones[0] / ones[1]


def load_checkpoint(cp, models, optimizer):
    for (m, c) in zip(models, cp['state_dict']):
        m.load_state_dict(c)
    return cp, models, optimizer, cp['loss'], cp['acc'], cp['epoch'], cp['best_acc']


def save_checkpoint(state_dict, isBest, save_path, model_name):
    if isBest:
        try:
            torch.save(state_dict, os.path.join(save_path, "best", model_name))
        except FileNotFoundError:
            os.mkdir(os.path.join(save_path, "best"))
            torch.save(state_dict, os.path.join(save_path, "best", model_name))
    else:
        try:
            torch.save(state_dict, os.path.join(save_path, "progress", model_name))
        except FileNotFoundError:
            os.mkdir(os.path.join(save_path, "progress"))
            torch.save(state_dict, os.path.join(save_path, "progress", model_name))


if __name__ == "__main__":
    os.environ["CUDA_VISIBLE_DEVICES"] = "8"
    filename = "/home/wine1865/다운로드/Ansan/data/Posturography_임상 검사 결과 데이터_결과판독_preliminary.xlsx"
    wb = load_workbook(filename)
    st = wb['Sheet1']
    valid_info = np.asarray(get_valid_index(st))
    train_idx = get_index("/home/wine1865/다운로드/Ansan/data/COP_train.txt")
    test_idx = get_index("/home/wine1865/다운로드/Ansan/data/COP_test.txt")
    all_idx = np.append(train_idx, test_idx)
    all_idx = valid_info[all_idx][:, 0]

    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    criterion = nn.BCEWithLogitsLoss()

    gt = 'rotary'
    split = 2000
    key = 'COG'
    sampling = 2
    scale_factor = 1
    mode = 'rec'
    gen_size = 200
    shuffle_trial = True

    max_epoch = 1000
    start_epoch = 0
    check_train = 10

    save_path = "/home/wine1865/다운로드/Ansan/models/11/BBN/cosine"

    for ci in range(1):
        # model settings
        model_shared = block_resnet_shared(in_channels=12)
        model_sep1 = block_resnet_separated(in_channels=256)
        model_sep2 = block_resnet_separated(in_channels=256)
        model_classifier1 = classifier(in_channels=512, num_classes=1)
        model_classifier2 = classifier(in_channels=512, num_classes=1)
        models = (model_shared, model_sep1, model_sep2, model_classifier1, model_classifier2)
        # model = nn.DataParallel(model)
        params = []
        for m in models:
            m.to(device)
            params += list(m.parameters())
        model_name = "BBN_BCEW_SOT2 6_SGD-1_R200_" + str(ci) + ".pth"
        print(toBlue("{}-th loop".format(ci)))
        # optimizer & scheduler
        optimizer = optim.SGD(params, lr=0.1, momentum=0.9, weight_decay=0.0001)
        scheduler = optim.lr_scheduler.CosineAnnealingWarmRestarts(optimizer, 100)
        # scheduler = optim.lr_scheduler.MultiStepLR(optimizer, [50, 150, 200, 300, 400], 0.1)
        # load previous model if it exists
        loss_toView = []
        acc_toView = []
        trainacc_toView = []
        best_acc = 0.0
        best_trainacc = 0.0
        if os.path.exists(os.path.join(save_path, "progress", model_name)):
            cp = torch.load(os.path.join(save_path, "progress", model_name))
            cp, models, optimizer, loss_toView, acc_toView, start_epoch, best_acc = \
                load_checkpoint(cp, models, optimizer)
            start_epoch += 1
            print(toGreen("Loaded previous model"),
                  '\t', "Avg loss: {}, acc: {}, lr: {}".format(
                    loss_toView[-1], acc_toView[-1], optimizer.param_groups[0]['lr']))
        # preparing data index
        train_info = []
        test_info = []
        for j, ind in enumerate(all_idx):
            step = int(len(all_idx) / 5)
            if (j >= step * ci) and (j < step * ci + step):
                test_info.append((ind, int(st.cell(ind, 1).value)))
            else:
                train_info.append((ind, int(st.cell(ind, 1).value)))
        # data loader
        sot = model_name.split("_")[2][3:].split(' ')
        print(sot)
        # testset = COP_Data(test_info, st, mode=mode, sot=sot, nperseg=gen_size,
        #                    rp_steps=gen_size, key=key, split=split, sampling=sampling, scale_factor=scale_factor,
        #                    shuffle_trial=shuffle_trial, gt=gt)
        trainset = COP_Data(train_info, st, mode=mode, sot=sot, nperseg=gen_size,
                            rp_steps=gen_size, key=key, split=split, sampling=sampling, scale_factor=scale_factor,
                            shuffle_trial=shuffle_trial, gt=gt)
        testset = COP_Data(test_info, st, mode=mode, sot=sot, nperseg=gen_size,
                           rp_steps=gen_size, key=key, split=split, sampling=sampling, scale_factor=scale_factor,
                           shuffle_trial=shuffle_trial, gt=gt)

        # todo: count sample numbers every fold loop
        sample_num_train = (293, 98)
        sample_num_test = (79, 18)

        batch_size = 8

        # epoch loop
        for e in range(start_epoch, max_epoch):
            if e > max_epoch - 1:
                break
            if (e + 1) % check_train == 0:
                with torch.no_grad():
                    train_acc, sot_acc_train, zero_acc, one_acc = \
                        test(e, models, trainset, device, sample_num_test, batch_size)
                trainacc_toView.append(train_acc)
                if train_acc > best_trainacc:
                    best_trainacc = train_acc
                print("Average: {}\tTotal: {}\tnormal: {}\tabnormal: {}"
                      .format(train_acc, sot_acc_train, zero_acc, one_acc))
            models, avg_loss = \
                train(e, max_epoch, models, trainset,
                      criterion, optimizer, scheduler, device, ci, sample_num_train, batch_size)
            # scheduler.step()
            loss_toView.append(avg_loss)
            print(toCyan("Epoch: {}\tAvg loss:{}\t".format(e, avg_loss)))
            with torch.no_grad():
                avg_acc, sot_acc, zero_acc, one_acc = test(e, models, testset, device, sample_num_test, batch_size)
            print(toCyan("Test average accuracy: {}\ttotal: {}\tnormal: {}\tabnormal: {}".format(
                avg_acc * 100.0, sot_acc, zero_acc, one_acc)))
            acc_toView.append(avg_acc)

            state_dict = []
            for m in models:
                state_dict.append(m.state_dict())
            save_checkpoint({
                'epoch': e,
                'state_dict': state_dict,
                'loss': loss_toView,
                'lr': optimizer.param_groups[0]['lr'],
                'acc': acc_toView,
                'best_acc': best_acc
            }, False, save_path, model_name)
            print(toGreen("Saved progress model"))
            if avg_acc > best_acc:
                best_acc = avg_acc
                save_checkpoint({
                    'epoch': e,
                    'state_dict': state_dict,
                    'loss': loss_toView,
                    'lr': optimizer.param_groups[0]['lr'],
                    'acc': acc_toView,
                    'best_acc': best_acc
                }, True, save_path, model_name)
                print(toGreen("Saved best model"))
        print(toGreen("BEST train acc: {}".format(best_trainacc)))
        print(toGreen("BEST test acc: {}".format(best_acc)))
    fig, axes = plt.subplots(1, 3)
    fig.suptitle("multitask1")
    axes[0].plot(loss_toView)
    axes[0].set_title("Loss")
    axes[1].plot(acc_toView)
    axes[1].set_title("Test Acc")
    axes[2].plot(trainacc_toView)
    axes[2].set_title("Train Acc")
    # axes[1, 0].plot(sot2_toView)
    # axes[1, 0].set_title("SOT2 ACC")
    # axes[1, 1].plot(sot4_toView)
    # axes[1, 1].set_title("SOT4 ACC")
    # axes[1, 2].plot(sot5_toView)
    # axes[1, 2].set_title("SOT5 ACC")
    plt.show()

